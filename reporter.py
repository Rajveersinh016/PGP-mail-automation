"""
reporter.py — PGP Container Glass Intelligence Platform
Generates a simplified, professional HTML email report with Market Pulse, metrics dashboard, and a single-sheet Excel report.
"""

import logging
import os
import tempfile
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# Badges and theme styling
CATEGORY_BADGES = {
    "🏭 New Container Glass Factory":  {"emoji": "🏭", "bg": "#d4edda", "color": "#155724"},
    "🏗 Plant Expansion":              {"emoji": "🏗", "bg": "#cce5ff", "color": "#004085"},
    "🔥 Furnace Rebuild":              {"emoji": "🔥", "bg": "#f8d7da", "color": "#721c24"},
    "⚙ Machinery & Equipment":         {"emoji": "⚙", "bg": "#e2d9f3", "color": "#432874"},
    "🤝 Customer Partnership":        {"emoji": "🤝", "bg": "#fde8d0", "color": "#7d3a00"},
    "💰 Investment":                  {"emoji": "💰", "bg": "#d2f4ea", "color": "#0f5132"},
    "🌱 Sustainability":              {"emoji": "🌱", "bg": "#d1e7dd", "color": "#0f5132"},
    "🍾 Beverage Industry":           {"emoji": "🍾", "bg": "#e2d9f3", "color": "#432874"},
    "🌸 Luxury Packaging":            {"emoji": "🌸", "bg": "#fce7f3", "color": "#9d174d"},
    "📦 Container Glass Packaging":   {"emoji": "📦", "bg": "#dbeafe", "color": "#1e40af"},
    "🔬 Technology":                  {"emoji": "🔬", "bg": "#cff4fc", "color": "#055160"},
    "🌍 Industry Update":             {"emoji": "🌍", "bg": "#e2e3e5", "color": "#383d41"},
}


def _build_html_email(articles: list[dict], report_date: str, market_pulse: str) -> str:
    """Build a clean, executive HTML email body."""
    total = len(articles)
    cards_html = ""

    # Calculate metrics
    companies_set = {a.get("company", "Unknown") for a in articles} - {"Unknown"}
    companies_count = len(companies_set)
    companies_str = ", ".join(sorted(companies_set)) if companies_set else "None"
    
    countries_set = {a.get("country", "Unknown") for a in articles} - {"Unknown"}
    countries_count = len(countries_set)
    countries_str = ", ".join(sorted(countries_set)) if countries_set else "None"

    # Category counts
    cat_counts = {}
    for a in articles:
        cat = a.get("category", "🌍 Industry Update")
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
    
    cat_breakdown_items = []
    for cat, count in sorted(cat_counts.items(), key=lambda x: x[1], reverse=True):
        emoji = CATEGORY_BADGES.get(cat, {}).get("emoji", "🌍")
        cat_breakdown_items.append(f"{emoji} {cat.replace(emoji, '').strip()}: {count}")
    cat_breakdown_str = " &nbsp;·&nbsp; ".join(cat_breakdown_items)

    for a in articles:
        cat = a.get("category", "🌍 Industry Update")
        badge = CATEGORY_BADGES.get(cat, CATEGORY_BADGES["🌍 Industry Update"])
        
        company = a.get("company", "Unknown")
        country = a.get("country", "Unknown")
        summary = a.get("summary", a.get("title", ""))
        key_details = a.get("key_details", "")
        source = a.get("source", "")
        pub_date = a.get("published", "")[:10] if a.get("published") else ""
        url = a.get("url", "#")
        business_impact = a.get("business_impact", "")
        priority = a.get("priority", "Low")
        confidence = a.get("confidence", "High")

        priority_color = "#dc2626" if priority == "High" else ("#d97706" if priority == "Medium" else "#4b5563")

        meta_html = ""
        if company != "Unknown":
            meta_html += f'<strong>Company:</strong> {company} &nbsp;|&nbsp; '
        if country != "Unknown":
            meta_html += f'<strong>Country:</strong> {country} &nbsp;|&nbsp; '
        if source:
            meta_html += f'<strong>Source:</strong> {source} &nbsp;|&nbsp; '
        if pub_date:
            meta_html += f'<strong>Date:</strong> {pub_date}'

        # Build Key Details HTML — convert bullet lines to styled list items
        key_details_html = ""
        if key_details:
            bullet_lines = [ln.strip().lstrip("•").strip() for ln in key_details.splitlines() if ln.strip()]
            if bullet_lines:
                items_html = "".join(
                    f'<li style="margin:0 0 4px 0; font-size:12px; color:#1e3a5f; line-height:1.45; font-family:Arial,sans-serif;">'
                    f'{line}</li>'
                    for line in bullet_lines
                )
                key_details_html = f"""
                  <div style="margin-top:8px; padding:10px 14px; background:#eef4ff; border-left:3px solid #3b82f6; border-radius:0 6px 6px 0;">
                    <p style="margin:0 0 5px 0; font-size:10px; font-weight:700; color:#1e40af; text-transform:uppercase; letter-spacing:0.5px; font-family:Arial,sans-serif;">🔍 Key Details</p>
                    <ul style="margin:0; padding-left:14px; list-style-type:disc;">{items_html}</ul>
                  </div>"""

        cards_html += f"""
        <tr>
          <td style="padding:16px 0; border-bottom:1px solid #e2e8f0;">
            <table width="100%" cellpadding="0" cellspacing="0" border="0">
              <tr>
                <td>
                  <!-- Category Badge & Priority -->
                  <div style="margin-bottom:6px;">
                    <span style="display:inline-block; background:{badge['bg']}; color:{badge['color']};
                                 font-size:10px; font-weight:700; padding:3px 8px; border-radius:12px;
                                 font-family:Arial,sans-serif; margin-right:6px;">
                      {badge['emoji']} {cat.upper()}
                    </span>
                    <span style="display:inline-block; font-size:10px; font-weight:700; color:{priority_color}; font-family:Arial,sans-serif; margin-right:6px;">
                      ⚠️ PRIORITY: {priority.upper()}
                    </span>
                    <span style="display:inline-block; font-size:10px; color:#475569; font-family:Arial,sans-serif;">
                      (Confidence: {confidence})
                    </span>
                  </div>
                  <!-- Title -->
                  <h3 style="margin:0 0 4px 0; font-size:15px; font-weight:700; color:#0f172a; line-height:1.4; font-family:Arial,sans-serif;">
                    <a href="{url}" style="color:#0f172a; text-decoration:none;" target="_blank">{a.get('title', '')}</a>
                  </h3>
                  <!-- Meta row -->
                  <p style="margin:0 0 8px 0; font-size:11px; color:#64748b; font-family:Arial,sans-serif;">
                    {meta_html}
                  </p>
                  <!-- Executive Summary block -->
                  <div style="background:#f8fafc; padding:12px 14px; border-left:3px solid #092f20; border-radius:0 6px 6px 0; margin-bottom:4px;">
                    <p style="margin:0 0 4px 0; font-size:10px; font-weight:700; color:#0f5132; text-transform:uppercase; letter-spacing:0.5px; font-family:Arial,sans-serif;">📋 Executive Summary</p>
                    <p style="margin:0; font-size:13px; color:#334155; line-height:1.5; font-family:Arial,sans-serif;">
                      {summary}
                    </p>
                  </div>
                  <!-- Key Details block -->
                  {key_details_html}
                  <!-- Business Impact -->
                  {f'''<p style="margin:8px 0 0 0; font-size:12px; color:#1e40af; line-height:1.5; font-family:Arial,sans-serif;">
                    <strong>💼 Business Impact:</strong> {business_impact}
                  </p>''' if business_impact else ''}
                  <!-- Link -->
                  <a href="{url}" target="_blank" style="display:inline-block; margin-top:8px; font-size:11.5px; font-weight:600; color:#092f20; text-decoration:none; font-family:Arial,sans-serif;">
                    Read Article →
                  </a>
                </td>
              </tr>
            </table>
          </td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Container Glass Intelligence Report</title>
</head>
<body style="margin:0; padding:0; background-color:#f1f5f9; font-family:Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#f1f5f9; padding:20px 0;">
    <tr>
      <td align="center">
        <table width="660" cellpadding="0" cellspacing="0" border="0" style="max-width:660px; background:#ffffff; border-radius:8px; overflow:hidden; box-shadow:0 1px 3px rgba(0,0,0,0.1);">
          <!-- Header -->
          <tr>
            <td style="background:#092f20; padding:24px 30px; border-bottom:3px solid #10b981;">
              <p style="margin:0 0 2px 0; font-size:10px; font-weight:600; color:#34d399; letter-spacing:1.5px; text-transform:uppercase;">
                GLOBAL CONTAINER GLASS MARKET INTELLIGENCE PLATFORM
              </p>
              <h1 style="margin:0; font-size:22px; font-weight:700; color:#ffffff; font-family:Arial,sans-serif;">
                Daily Executive Report
              </h1>
              <p style="margin:4px 0 0 0; font-size:12px; color:#93c5fd; font-family:Arial,sans-serif;">
                {report_date} &nbsp;·&nbsp; {total} Curated Industry Updates
              </p>
            </td>
          </tr>
          <!-- Body -->
          <tr>
            <td style="padding:20px 30px 30px 30px;">
              <!-- Market Pulse Paragraph -->
              <div style="background:#f0fdf4; border-left:4px solid #10b981; padding:15px; margin-bottom:20px; border-radius:4px;">
                <h3 style="margin:0 0 6px 0; color:#0f2f20; font-family:Arial,sans-serif; font-size:13.5px; font-weight:700; letter-spacing:0.5px; text-transform:uppercase;">📈 Market Pulse</h3>
                <p style="margin:0; font-size:13px; color:#1e293b; line-height:1.55; font-family:Arial,sans-serif;">
                  {market_pulse}
                </p>
              </div>

              <!-- Pulse Stats Dashboard -->
              <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-bottom:15px;">
                <tr>
                  <td width="32%" align="center" style="background:#f8fafc; padding:14px; border-radius:6px; border:1px solid #e2e8f0;">
                    <span style="font-size:24px; font-weight:700; color:#092f20;">{total}</span><br>
                    <span style="font-size:9.5px; color:#64748b; font-weight:700; text-transform:uppercase; letter-spacing:0.5px;">Relevant News</span>
                  </td>
                  <td width="2%"></td>
                  <td width="32%" align="center" style="background:#f8fafc; padding:14px; border-radius:6px; border:1px solid #e2e8f0;">
                    <span style="font-size:24px; font-weight:700; color:#092f20;">{companies_count}</span><br>
                    <span style="font-size:9.5px; color:#64748b; font-weight:700; text-transform:uppercase; letter-spacing:0.5px;">Companies</span>
                  </td>
                  <td width="2%"></td>
                  <td width="32%" align="center" style="background:#f8fafc; padding:14px; border-radius:6px; border:1px solid #e2e8f0;">
                    <span style="font-size:24px; font-weight:700; color:#092f20;">{countries_count}</span><br>
                    <span style="font-size:9.5px; color:#64748b; font-weight:700; text-transform:uppercase; letter-spacing:0.5px;">Countries</span>
                  </td>
                </tr>
              </table>

              <!-- Dashboard Details -->
              <div style="background:#f8fafc; padding:12px 15px; border-radius:6px; border:1px solid #e2e8f0; margin-bottom:20px; font-size:11.5px; font-family:Arial,sans-serif; color:#334155; line-height:1.6;">
                <div style="margin-bottom:4px;"><strong>🏢 Monitored Companies:</strong> {companies_str}</div>
                <div style="margin-bottom:4px;"><strong>🌍 Geographies Impacted:</strong> {countries_str}</div>
                <div><strong>📊 Category Spans:</strong> <span style="color:#475569;">{cat_breakdown_str}</span></div>
              </div>

              <!-- Curated Articles Divider -->
              <h2 style="font-size:15px; font-weight:700; color:#0f172a; margin:25px 0 10px 0; border-bottom:2px solid #092f20; padding-bottom:6px; text-transform:uppercase; letter-spacing:0.5px; font-family:Arial,sans-serif;">
                Curated Updates
              </h2>

              <!-- Article Cards -->
              <table width="100%" cellpadding="0" cellspacing="0" border="0">
                {cards_html}
              </table>
            </td>
          </tr>
          <!-- Footer -->
          <tr>
            <td style="background:#f8fafc; border-top:1px solid #e2e8f0; padding:16px 30px; font-size:11px; color:#64748b; line-height:1.5;">
              This enterprise intelligence report is automatically generated for internal strategic analysis at PGP Glass.<br>
              The complete structured database is attached as a single-sheet Excel workbook.
            </td>
          </tr>
        </table>
      </td>
    </tr>
  </table>
</body>
</html>"""
    return html


def _build_excel(articles: list[dict]) -> str:
    """Build a single-sheet Excel workbook containing all metadata fields. Returns file path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Curated Articles"
    ws.row_dimensions[1].height = 25

    headers = [
        "Company", "Country", "Category", "Executive Summary", "Key Details",
        "Business Impact", "Priority", "Confidence", "Source", "Published Date", "URL"
    ]
    ws.append(headers)

    # Style Header Row
    fill = PatternFill(start_color="092F20", end_color="092F20", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(bottom=Side(style="medium", color="10B981"))

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border

    alt_fill = PatternFill(start_color="F4Fbf7", end_color="F4Fbf7", fill_type="solid")
    normal_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for i, a in enumerate(articles, 1):
        row_fill = alt_fill if i % 2 == 0 else normal_fill
        pub_date = a.get("published", "")[:10] if a.get("published") else ""
        row = [
            a.get("company", "Unknown"),
            a.get("country", "Unknown"),
            a.get("category", "🌍 Industry Update"),
            a.get("summary", ""),
            a.get("key_details", ""),
            a.get("business_impact", ""),
            a.get("priority", "Low"),
            a.get("confidence", "High"),
            a.get("source", ""),
            pub_date,
            a.get("url", ""),
        ]
        ws.append(row)
        
        # Apply Row styling
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=i + 1, column=col_idx)
            cell.fill = row_fill
            cell.font = Font(name="Calibri", size=9)
            # Wrap text for summaries, key details, impacts, and URLs
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in (4, 5, 6, 11)))

    # Set column widths (added Key Details column)
    widths = [18, 14, 24, 50, 50, 45, 12, 12, 18, 12, 25]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Save to local temp directory relative to the project directory
    project_dir = os.path.dirname(os.path.abspath(__file__))
    temp_dir = os.path.join(project_dir, "temp")
    os.makedirs(temp_dir, exist_ok=True)

    date_str = datetime.now(timezone.utc).strftime("%Y%m%d")
    filename = f"container_glass_report_{date_str}.xlsx"
    excel_path = os.path.join(temp_dir, filename)
    wb.save(excel_path)
    return excel_path


def generate_report(articles: list[dict], market_pulse: str) -> tuple[str, str]:
    """Public interface: returns HTML email body and Excel file path."""
    report_date = datetime.now(timezone.utc).strftime("%B %d, %Y")
    
    # Sort articles by priority (High -> Medium -> Low), then by Company name
    priority_order = {"High": 1, "Medium": 2, "Low": 3}
    articles = sorted(articles, key=lambda x: (
        priority_order.get(x.get("priority", "Low"), 3),
        x.get("company", "Unknown").lower()
    ))

    html_body = _build_html_email(articles, report_date, market_pulse)
    excel_path = _build_excel(articles)

    return html_body, excel_path
